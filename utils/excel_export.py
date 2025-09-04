import sys
import os
import pandas as pd
from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Safe path handling
try:
    current_dir = Path(__file__).parent
    sys.path.append(str(current_dir.parent))
except NameError:
    sys.path.append('..')

class ExcelExportService:
    """Fixed Excel export service for admin review reports"""
    
    def __init__(self, export_directory: str = "data"):
        self.export_directory = export_directory
        self.ensure_export_directory()
        self.export_history = []
        
        # Excel file paths
        self.appointments_file = os.path.join(export_directory, "appointments.xlsx")
        self.admin_report_file = os.path.join(export_directory, "admin_review_report.xlsx")
        self.daily_summary_file = os.path.join(export_directory, "daily_summary.xlsx")
    
    def ensure_export_directory(self):
        """Ensure export directory exists"""
        os.makedirs(self.export_directory, exist_ok=True)
    
    def export_appointment_data(self, appointment_data: Dict[str, Any]) -> Tuple[str, bool]:
        """
        Export single appointment data to Excel (assignment requirement)
        
        Args:
            appointment_data: Complete appointment information
            
        Returns:
            - response_message: Export result message
            - success: Whether export was successful
        """
        
        try:
            # Prepare data for export
            export_row = self._prepare_appointment_row(appointment_data)
            
            # Load existing data or create new
            if os.path.exists(self.appointments_file):
                df = pd.read_excel(self.appointments_file)
            else:
                df = pd.DataFrame(columns=self._get_appointment_columns())
            
            # Add new appointment data
            df = pd.concat([df, pd.DataFrame([export_row])], ignore_index=True)
            
            # Export to Excel with formatting
            success = self._export_dataframe_to_excel(
                df, 
                self.appointments_file, 
                "Appointments",
                "appointment_data"
            )
            
            if success:
                # Log export
                self._log_export("appointment_data", appointment_data.get('appointment_id', 'Unknown'))
                
                patient_name = self._extract_patient_name(appointment_data)
                
                response_message = (
                    f"📊 **Appointment Data Exported Successfully!**\n\n"
                    f"**File**: {self.appointments_file}\n"
                    f"**Appointment ID**: {appointment_data.get('appointment_id', 'N/A')}\n"
                    f"**Patient**: {patient_name}\n"
                    f"**Export Time**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    f"**Data Exported**:\n"
                    f"• Patient Information\n"
                    f"• Appointment Details\n"
                    f"• Insurance Information\n"
                    f"• Confirmation Status\n\n"
                    f"**Admin Review**: Ready for administrative review"
                )
                
                return response_message, True
            else:
                return "❌ Failed to export appointment data to Excel.", False
                
        except Exception as e:
            print(f"❌ Error exporting appointment data: {e}")
            return f"❌ Error exporting appointment data: {str(e)}", False
    
    def _extract_patient_name(self, appointment_data: Dict[str, Any]) -> str:
        """Extract patient name from various data structures"""
        # Try direct access first
        if appointment_data.get('patient_name'):
            return appointment_data['patient_name']
        
        # Try nested patient_info
        patient_info = appointment_data.get('patient_info', {})
        if patient_info.get('patient_name'):
            return patient_info['patient_name']
        
        return 'N/A'
    
    def _prepare_appointment_row(self, appointment_data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare single appointment data for export with proper data extraction"""
        
        # Handle nested data structures from confirmation agent
        patient_info = appointment_data.get('patient_info', {})
        appointment_slot = appointment_data.get('appointment_slot', {})
        insurance_info = appointment_data.get('insurance_info', {})
        
        # Create the export row with proper data extraction
        export_row = {
            'appointment_id': appointment_data.get('appointment_id', ''),
            'patient_id': appointment_data.get('patient_id', ''),
            
            # Patient information - handle both nested and direct access
            'patient_name': (
                appointment_data.get('patient_name') or 
                patient_info.get('patient_name', '')
            ),
            'patient_email': (
                appointment_data.get('patient_email') or 
                patient_info.get('email', '')
            ),
            'patient_phone': (
                appointment_data.get('patient_phone') or 
                patient_info.get('phone', '')
            ),
            'date_of_birth': (
                appointment_data.get('date_of_birth') or 
                patient_info.get('date_of_birth', '')
            ),
            
            # Appointment slot information - handle both nested and direct access
            'doctor': (
                appointment_data.get('doctor') or 
                appointment_slot.get('doctor', '')
            ),
            'location': (
                appointment_data.get('location') or 
                appointment_slot.get('location', '')
            ),
            'appointment_date': (
                appointment_data.get('appointment_date') or 
                appointment_slot.get('date', '')
            ),
            'appointment_time': (
                appointment_data.get('appointment_time') or 
                appointment_slot.get('time', '')
            ),
            'duration': (
                appointment_data.get('duration') or 
                appointment_slot.get('duration_available', '')
            ),
            
            # Insurance information - handle both nested and direct access
            'insurance_carrier': (
                appointment_data.get('insurance_carrier') or 
                insurance_info.get('primary_carrier', '')
            ),
            'member_id': (
                appointment_data.get('member_id') or 
                insurance_info.get('member_id', '')
            ),
            'group_number': (
                appointment_data.get('group_number') or 
                insurance_info.get('group_number', '')
            ),
            
            # Status and timestamps
            'status': appointment_data.get('status', ''),
            'confirmed_at': appointment_data.get('confirmed_at', ''),
            'created_at': appointment_data.get('created_at', ''),
            'reminders_sent': appointment_data.get('reminders_sent', 0),
            'form_sent': appointment_data.get('form_sent', False),
            'excel_exported': True,
            'exported_at': datetime.now().isoformat()
        }
        
        return export_row
    
    def generate_admin_review_report(self, appointments_data: List[Dict[str, Any]]) -> Tuple[str, bool]:
        """
        Generate comprehensive admin review report (assignment requirement)
        """
        
        try:
            if not appointments_data:
                return "⚠️ No appointment data available for admin report.", False
            
            # Create comprehensive report
            report_data = self._prepare_admin_report_data(appointments_data)
            
            # Export to Excel with professional formatting
            success = self._export_dataframe_to_excel(
                report_data,
                self.admin_report_file,
                "Admin Review Report",
                "admin_report"
            )
            
            if success:
                # Log export
                self._log_export("admin_report", f"{len(appointments_data)} appointments")
                
                response_message = (
                    f"📊 **Admin Review Report Generated Successfully!**\n\n"
                    f"**File**: {self.admin_report_file}\n"
                    f"**Total Appointments**: {len(appointments_data)}\n"
                    f"**Report Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    f"**Report Contents**:\n"
                    f"• Appointment Summary\n"
                    f"• Patient Demographics\n"
                    f"• Insurance Analysis\n"
                    f"• Scheduling Patterns\n"
                    f"• Revenue Overview\n\n"
                    f"**Ready for**: Administrative review and decision making"
                )
                
                return response_message, True
            else:
                return "❌ Failed to generate admin review report.", False
                
        except Exception as e:
            print(f"❌ Error generating admin report: {e}")
            return f"❌ Error generating admin report: {str(e)}", False
    
    def _prepare_admin_report_data(self, appointments_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare comprehensive admin report data"""
        
        # Convert to DataFrame for easier analysis
        processed_data = []
        for appointment in appointments_data:
            processed_row = self._prepare_appointment_row(appointment)
            processed_data.append(processed_row)
        
        df = pd.DataFrame(processed_data)
        
        # Create summary statistics
        summary_data = []
        
        # Patient demographics
        total_patients = len(df)
        
        summary_data.append({
            'metric': 'Total Appointments',
            'value': total_patients,
            'category': 'Overview'
        })
        
        # Doctor distribution
        if 'doctor' in df.columns and not df['doctor'].isna().all():
            doctor_counts = df['doctor'].value_counts().to_dict()
            for doctor, count in doctor_counts.items():
                if pd.notna(doctor) and doctor:
                    summary_data.append({
                        'metric': f'Appointments with {doctor}',
                        'value': count,
                        'category': 'Doctor Distribution'
                    })
        
        # Location distribution
        if 'location' in df.columns and not df['location'].isna().all():
            location_counts = df['location'].value_counts().to_dict()
            for location, count in location_counts.items():
                if pd.notna(location) and location:
                    summary_data.append({
                        'metric': f'Appointments at {location}',
                        'value': count,
                        'category': 'Location Distribution'
                    })
        
        # Insurance analysis
        if 'insurance_carrier' in df.columns and not df['insurance_carrier'].isna().all():
            insurance_counts = df['insurance_carrier'].value_counts().to_dict()
            for carrier, count in insurance_counts.items():
                if pd.notna(carrier) and carrier:
                    summary_data.append({
                        'metric': f'Patients with {carrier}',
                        'value': count,
                        'category': 'Insurance Analysis'
                    })
        
        # Revenue estimation
        if 'duration' in df.columns and not df['duration'].isna().all():
            total_minutes = df['duration'].sum()
            total_revenue = total_minutes * 2  # $2 per minute estimate
            summary_data.append({
                'metric': 'Estimated Revenue',
                'value': f"${total_revenue:,.2f}",
                'category': 'Financial'
            })
        
        return pd.DataFrame(summary_data)
    
    def _get_appointment_columns(self) -> List[str]:
        """Get standard appointment export columns"""
        return [
            'appointment_id', 'patient_id', 'patient_name', 'patient_email', 'patient_phone',
            'date_of_birth', 'doctor', 'location', 'appointment_date',
            'appointment_time', 'duration', 'insurance_carrier', 'member_id',
            'group_number', 'status', 'confirmed_at', 'created_at', 'reminders_sent', 
            'form_sent', 'excel_exported', 'exported_at'
        ]
    
    def _export_dataframe_to_excel(self, 
                                  df: pd.DataFrame, 
                                  file_path: str, 
                                  sheet_name: str,
                                  export_type: str) -> bool:
        """Export DataFrame to Excel with professional formatting"""
        
        try:
            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply professional formatting
                self._apply_excel_formatting(worksheet, export_type)
                
            return True
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
            return False
    
    def _apply_excel_formatting(self, worksheet, export_type: str):
        """Apply professional Excel formatting"""
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _log_export(self, export_type: str, identifier: str):
        """Log export activity"""
        export_record = {
            'export_type': export_type,
            'identifier': identifier,
            'timestamp': datetime.now().isoformat(),
            'file_path': self.export_directory
        }
        self.export_history.append(export_record)
    
    def get_export_summary(self) -> str:
        """Get summary of all exports"""
        
        if not self.export_history:
            return "📊 **Export Summary**: No exports performed yet."
        
        summary = f"📊 **Excel Export Summary**\n\n"
        summary += f"**Total Exports**: {len(self.export_history)}\n\n"
        
        # Group by export type
        export_types = {}
        for export in self.export_history:
            export_type = export['export_type']
            if export_type not in export_types:
                export_types[export_type] = 0
            export_types[export_type] += 1
        
        summary += "**Export Types**:\n"
        for export_type, count in export_types.items():
            summary += f"• {export_type}: {count}\n"
        
        summary += "\n**Recent Exports**:\n"
        for export in self.export_history[-5:]:  # Show last 5
            summary += (
                f"• {export['export_type']} - {export['identifier']} - "
                f"{export['timestamp'][:10]}\n"
            )
        
        return summary
    
    def clear_export_history(self):
        """Clear export history"""
        self.export_history = []

# Test function
def test_fixed_excel_export():
    """Test the fixed Excel export service"""
    print("🧪 Testing Fixed Excel Export Service...\n")
    
    service = ExcelExportService()
    
    # Test data with nested structure (like from confirmation agent)
    test_appointment = {
        'appointment_id': 'APT_20250903_FIXED_001',
        'patient_id': 'PAT_FIXED_001',
        'patient_info': {
            'patient_name': 'Alice Johnson',
            'email': 'alice.johnson@test.com',
            'phone': '(555) 111-2222',
            'date_of_birth': '03/15/1985'
        },
        'appointment_slot': {
            'doctor': 'Dr. Aish',
            'location': 'Banjara Hills',
            'date': '2025-01-17',
            'time': '11:00',
            'duration_available': 30
        },
        'insurance_info': {
            'primary_carrier': 'Cigna',
            'member_id': 'CG555666777',
            'group_number': 'GRP003'
        },
        'status': 'confirmed',
        'confirmed_at': datetime.now().isoformat(),
        'created_at': datetime.now().isoformat(),
        'reminders_sent': 0,
        'form_sent': False
    }
    
    print("=== Testing Fixed Single Appointment Export ===")
    response, success = service.export_appointment_data(test_appointment)
    print(f"Export Result: {'✅ Success' if success else '❌ Failed'}")
    print(f"Response: {response}")
    
    if success:
        print("\n=== Testing Fixed Admin Report Generation ===")
        admin_response, admin_success = service.generate_admin_review_report([test_appointment])
        print(f"Admin Report Result: {'✅ Success' if admin_success else '❌ Failed'}")
        print(f"Admin Response: {admin_response}")
    
    print(f"\n{service.get_export_summary()}")

if __name__ == "__main__":
    test_fixed_excel_export()